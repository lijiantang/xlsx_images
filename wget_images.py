#! /usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8

"""
下载xlsx图片
"""
from openpyxl.reader.excel import load_workbook
import re
import sys
import datetime
import os
import time
import urllib.request


#取第一张表
table = sys.argv[1]
wb = load_workbook('%s.xlsx' % table)
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])

print ("表名:",ws.title)
print ("总行数:", ws.max_row)
print ("总列数:",ws.max_column)

####创建文件夹
today = datetime.datetime.now()
todayStr = today.strftime("%Y%m%d%H%M")
if not os.path.exists(todayStr):
    os.mkdir(todayStr)

def get_image(url,path):
    opener=urllib.request.build_opener()
    opener.addheaders=[('User-Agent','Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.1941.0 Safari/537.36')]
    urllib.request.install_opener(opener)
    urllib.request.urlretrieve(url,path)


for rx in range(2, ws.max_row+1):
    keys = ws.cell(row=rx, column=1).value
    values = ws.cell(row=rx, column=2).value

    url = str(values)

    if url == 'None':
        continue

    strinfo = re.compile('60x60')
    newurl = strinfo.sub('600x600',url)

    jpg_name = todayStr + '/' + str(keys) + '.jpg'
    get_image(newurl,jpg_name)
    print (jpg_name)
    time.sleep (1)

