#! /usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8
#
# Copyright © 2018 Jerry <mr.jerry.li@letote.cn>
#
# Distributed under terms of the MIT license.

"""
批量改名
"""
from openpyxl.reader.excel import load_workbook
import sys
import os


#取第一张表
table = sys.argv[1]
wb = load_workbook('%s.xlsx' % table)
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])

print ("表名:",ws.title)
print ("总行数:", ws.max_row)
print ("总列数:",ws.max_column)

#建立字典存储表格内容
data_dic = {}

for rx in range(2, ws.max_row+1):
    pid = rx - 1
    w1 = ws.cell(row=rx, column=2).value
    data_dic[pid] = w1


for keys,values in data_dic.items():
   # print (keys,values)
    if values == 'None':
        continue
    pic_name = 'media/' + str(values) + '.png'
    old_name = 'media/image' + str(keys) + '.png'
    os.rename(old_name,pic_name)
    print (old_name,"------->",pic_name)

